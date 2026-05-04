import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────
GREEN  = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
RED    = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")
YELLOW = PatternFill(start_color="FFF176", end_color="FFF176", fill_type="solid")
HEADER = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")

def _col_index(ws, header: str) -> str:
    """Return the Excel column letter for a given header name."""
    for cell in ws[1]:
        if cell.value == header:
            return get_column_letter(cell.column)
    raise ValueError(f"Column '{header}' not found in worksheet headers.")

def generate_excel(df: pd.DataFrame, filename: str = "price_report.xlsx") -> None:
    # ── Write raw data ─────────────────────────────────────────
    df.to_excel(filename, index=False)

    wb = load_workbook(filename)
    ws = wb.active

    # ── Locate columns dynamically (safe against reordering) ───
    col_mytek_price      = _col_index(ws, "mytek_price")
    col_tunisianet_price = _col_index(ws, "tunisianet_price")
    col_cheapest         = _col_index(ws, "cheapest_source")

    # ── Style header row ───────────────────────────────────────
    for cell in ws[1]:
        cell.fill      = HEADER
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # ── Colour price cells per row ─────────────────────────────
    for row in range(2, ws.max_row + 1):
        source = ws[f"{col_cheapest}{row}"].value

        if source == "Mytek":
            ws[f"{col_mytek_price}{row}"].fill      = GREEN
            ws[f"{col_tunisianet_price}{row}"].fill = RED
        elif source == "Tunisianet":
            ws[f"{col_tunisianet_price}{row}"].fill = GREEN
            ws[f"{col_mytek_price}{row}"].fill      = RED
        else:  # Same
            ws[f"{col_mytek_price}{row}"].fill      = YELLOW
            ws[f"{col_tunisianet_price}{row}"].fill = YELLOW

    # ── Auto-fit column widths ─────────────────────────────────
    for col in ws.columns:
        max_len    = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    wb.save(filename)
    print(f"📊 Excel report saved → {filename}")